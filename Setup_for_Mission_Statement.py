# file:  Setup_Mission_Statement.py
#  Create Mission_Statement_Workbook
#
# Revision History
#    DATE          AUTHOR         DESCRIPTION
#  12/15/2013     TEV           Added the openpyxl code to
#                                        create a workbook and workseets
#                                        tabs.  Left the math, sys libraries.
#                                        Added numpy, scipy and rpy2 libraries
import xlwt
import numpy
import scipy
import math
import rpy2
import scp
import sys
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter

# Create a workbook in memory
#That has only one worksheet/spreadsheet
wb = Workbook()
from openpyxl.cell import get_column_letter


ws = wb.get_active_sheet()
ws1 = wb.create_sheet(title="Contracts_List") 
ws1a = wb.create_sheet(title="Contractor_List")
ws3 = wb.create_sheet(title="Product_List")
ws3a = wb.create_sheet(title="Accounts")
ws4=  wb.create_sheet(title="Product1_BOM")
ws5 = wb.create_sheet(title="Product1_WIP")
ws5a = wb.create_sheet(title="Product1_AR")
ws5b = wb.create_sheet(title="Product1_AP")
ws5c = wb.create_sheet(title="Product1_GA")

                       

ws6 = wb.create_sheet(title="Grants")
ws7 = wb.create_sheet(title="SOW")
ws8 = wb.create_sheet(title="Documenentation")
ws9 = wb.create_sheet(title="Gov_Supplied_HW")
ws10 = wb.create_sheet(title="Gov_Supplied_SW")



ws2 = wb.get_sheet_by_name("Vendor_List")
ws is ws2


print wb.get_sheet_names()



file_name = '/working/log.txt'

        
try:
    with open(file_name, "w") as outfile:
        outfile.write('\n\t\tReplace me with a spreadsheet\n\n')
except IOError:
    print 'oops!'
            